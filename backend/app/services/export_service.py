from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def generar_excel(titulo: str, columnas: list[str], filas: list[list]) -> BytesIO:
    """Genera un archivo Excel con titulo, headers y datos.

    Returns:
        BytesIO listo para StreamingResponse.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Titulo en fila 1 (merge cells)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
    cell_titulo = ws.cell(row=1, column=1, value=titulo)
    cell_titulo.font = Font(bold=True, size=14)
    cell_titulo.alignment = Alignment(horizontal="center")

    # Headers en fila 3
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Datos desde fila 4
    for row_idx, fila in enumerate(filas, 4):
        for col_idx, valor in enumerate(fila, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    # Auto-width columnas
    for col_idx in range(1, len(columnas) + 1):
        max_length = len(str(columnas[col_idx - 1]))
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
